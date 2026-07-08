from __future__ import annotations

import csv
import json
import re
import shutil
from io import BytesIO, StringIO
from pathlib import Path

from cryptography.exceptions import InvalidTag

from .config import OUTPUTS_DIR, STUDENTS_DIR, ensure_data_dirs
from .data_crypto import ENC_MARKER
from .io_utils import read_table_file
from .models import StudentInput, new_id
from .secure_io import load_secure_json, save_secure_json
from .storage_policy import student_dict_for_disk
from .write_sections import WRITE_SECTIONS


def _student_path(student_id: str) -> Path:
    return STUDENTS_DIR / f"{student_id}.json"


def _read_file_data(path: Path) -> dict | None:
    try:
        data = load_secure_json(path)
    except (OSError, ValueError, json.JSONDecodeError, RuntimeError, InvalidTag):
        return None
    if isinstance(data, dict) and data.get(ENC_MARKER):
        return None
    if not isinstance(data, dict):
        return {}
    return data


def _load_student_from_path(path: Path) -> StudentInput | None:
    data = _read_file_data(path)
    if data is None:
        return None
    return StudentInput.from_dict({**data, "id": path.stem})


def _resolve_student_path(student_id: str) -> Path | None:
    direct = _student_path(student_id)
    if direct.exists():
        return direct
    ensure_data_dirs()
    for path in STUDENTS_DIR.glob("*.json"):
        if path.stem == student_id:
            return path
        data = _read_file_data(path)
        if data is None:
            continue
        if str(data.get("id", "")).strip() == student_id:
            return path
    return None


def _student_has_content(student: StudentInput) -> bool:
    if student.name.strip():
        return True
    if student.generated:
        return True
    notes = student.notes or {}
    if str(notes.get("행발") or notes.get("행발_notes") or "").strip():
        return True
    if notes.get("keywords"):
        return True
    if notes.get("write_targets"):
        return True
    if student.subjects:
        return True
    if any(str(value or "").strip() for value in (student.changche or {}).values()):
        return True
    return False


def _is_ghost_student(student: StudentInput) -> bool:
    return not _student_has_content(student)


def _ensure_student_id_matches_file(student: StudentInput, path: Path) -> None:
    if student.id == path.stem:
        return
    old_id = student.id
    student.id = path.stem
    save_student(student)
    stale = _student_path(old_id)
    if stale.exists() and stale.resolve() != path.resolve():
        stale.unlink(missing_ok=True)


def reconcile_students(*, remove_ghosts: bool = True) -> dict[str, list[str]]:
    ensure_data_dirs()
    removed: list[str] = []
    fixed: list[str] = []
    for path in sorted(STUDENTS_DIR.glob("*.json")):
        student = _load_student_from_path(path)
        if student is None:
            continue
        if remove_ghosts and _is_ghost_student(student):
            path.unlink(missing_ok=True)
            removed.append(path.name)
            continue
        data = _read_file_data(path)
        if isinstance(data, dict) and data.get("id") != path.stem:
            _ensure_student_id_matches_file(student, path)
            fixed.append(path.stem)
    return {"removed": removed, "fixed": fixed}


def list_students(*, status: str | None = None) -> list[StudentInput]:
    reconcile_students()
    ensure_data_dirs()
    students: list[StudentInput] = []
    for path in sorted(STUDENTS_DIR.glob("*.json")):
        student = _load_student_from_path(path)
        if student is None:
            continue
        if status is None or student.status == status:
            students.append(student)
    students.sort(key=lambda s: (s.grade, s.class_num, s.number, s.name))
    return students


def get_student(student_id: str) -> StudentInput | None:
    path = _resolve_student_path(student_id)
    if not path:
        return None
    student = _load_student_from_path(path)
    if student is None:
        return None
    _ensure_student_id_matches_file(student, path)
    return student


def save_student(student: StudentInput) -> StudentInput:
    ensure_data_dirs()
    save_secure_json(_student_path(student.id), student_dict_for_disk(student))
    return student


def add_student(student: StudentInput) -> StudentInput:
    if not student.id:
        student.id = new_id()
    return save_student(student)


def _student_output_dir(student_id: str) -> Path:
    return OUTPUTS_DIR / student_id


def delete_student(student_id: str) -> bool:
    path = _resolve_student_path(student_id)
    if not path:
        return False
    file_id = path.stem
    path.unlink()
    output_dir = _student_output_dir(file_id)
    if output_dir.exists():
        shutil.rmtree(output_dir, ignore_errors=True)
    stale = _student_path(student_id)
    if stale.exists() and stale.resolve() != path.resolve():
        stale.unlink(missing_ok=True)
    return True


def delete_students(student_ids: list[str]) -> dict[str, list[str] | int]:
    deleted: list[str] = []
    not_found: list[str] = []
    for student_id in student_ids:
        if delete_student(student_id):
            deleted.append(student_id)
        else:
            not_found.append(student_id)
    return {"deleted": deleted, "not_found": not_found, "count": len(deleted)}


def delete_all_students() -> int:
    ensure_data_dirs()
    paths = list(STUDENTS_DIR.glob("*.json"))
    for path in paths:
        file_id = path.stem
        path.unlink(missing_ok=True)
        output_dir = _student_output_dir(file_id)
        if output_dir.exists():
            shutil.rmtree(output_dir, ignore_errors=True)
    return len(paths)


def reset_student_generated(student: StudentInput) -> StudentInput:
    student.status = "pending"
    student.generated = {}
    student.error_message = ""
    return save_student(student)


def reset_generated_for_student(student_id: str) -> bool:
    student = get_student(student_id)
    if not student:
        return False
    if not student.generated:
        return False
    reset_student_generated(student)
    return True


def reset_generated_for_students(student_ids: list[str]) -> dict[str, list[str] | int]:
    reset: list[str] = []
    not_found: list[str] = []
    for student_id in student_ids:
        student = get_student(student_id)
        if not student:
            not_found.append(student_id)
            continue
        if student.generated:
            reset_student_generated(student)
            reset.append(student_id)
    return {"reset": reset, "not_found": not_found, "count": len(reset)}


def reset_all_generated() -> int:
    count = 0
    for student in list_students():
        if student.generated:
            reset_student_generated(student)
            count += 1
    return count


def export_students_xlsx(students: list[StudentInput] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = students if students is not None else list_students()
    subjects: list[str] = []
    seen_subjects: set[str] = set()
    for student in rows:
        for subject in (student.generated.get("세특") or {}).keys():
            if subject not in seen_subjects:
                seen_subjects.add(subject)
                subjects.append(subject)

    headers = ["학년", "반", "번호", "이름", "상태", "행발", *[
        f"세특_{subject}" for subject in subjects
    ], "자율", "동아리", "봉사", "진로"]

    wb = Workbook()
    ws = wb.active
    ws.title = "생기부"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for student in rows:
        generated = student.generated or {}
        setuk = generated.get("세특") or {}
        changche = generated.get("창체") or {}
        ws.append([
            student.grade,
            student.class_num,
            student.number,
            student.name,
            student.status,
            str(generated.get("행발") or ""),
            *[str(setuk.get(subject) or "") for subject in subjects],
            str(changche.get("자율") or ""),
            str(changche.get("동아리") or ""),
            str(changche.get("봉사") or ""),
            str(changche.get("진로") or ""),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _split_multi(value: str) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[;|/]", value)
    return [p.strip() for p in parts if p.strip()]


REGISTRY_CHANGCHE_KEYS = ("자율", "동아리", "진로")
REGISTRY_WRITE_TARGET_KEYS = ("write_targets", "작성대상", "작성_대상")
REGISTRY_SUBJECT_SUFFIXES: tuple[tuple[str, str], ...] = (
    ("_진로", "career"),
    ("_수행평가", "assessment_type"),
    ("_주제", "topic"),
    ("_활동", "content"),
)


def _row_int(row: dict[str, str], *keys: str, default: int = 0) -> int:
    for key in keys:
        raw = row.get(key, "").strip()
        if raw:
            try:
                return int(raw)
            except ValueError:
                return default
    return default


def _row_str(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key, "").strip()
        if value:
            return value
    return ""


def _row_has_key(row: dict[str, str], *keys: str) -> bool:
    return any(key in row for key in keys)


def _write_targets_from_row(row: dict[str, str]) -> list[str]:
    raw = _row_str(row, *REGISTRY_WRITE_TARGET_KEYS)
    if not raw:
        return []
    allowed = set(WRITE_SECTIONS)
    targets: list[str] = []
    for item in _split_multi(raw):
        if item in allowed and item not in targets:
            targets.append(item)
    return targets


def _subjects_present_in_row(row: dict[str, str]) -> set[str]:
    subjects: set[str] = set()
    for key in row.keys():
        if key.startswith("세특_"):
            rest = key[3:]
            for suffix, _field in REGISTRY_SUBJECT_SUFFIXES:
                if rest.endswith(suffix):
                    subject = rest[: -len(suffix)]
                    if subject:
                        subjects.add(subject)
                    break
            else:
                if rest:
                    subjects.add(rest)
        elif key.startswith("subject_"):
            subject = key.split("_", 1)[1]
            if subject:
                subjects.add(subject)
        elif key.endswith("_activities"):
            subject = key.removesuffix("_activities")
            if subject:
                subjects.add(subject)
    return subjects


def student_identity_key(
    *,
    name: str,
    grade: int,
    class_num: int,
    number: int,
) -> tuple[int, int, int, str]:
    return (grade, class_num, number, name.strip())


def find_existing_for_row(row: dict[str, str]) -> StudentInput | None:
    student_id = _row_str(row, "id")
    if student_id:
        existing = get_student(student_id)
        if existing:
            return existing

    name = _row_str(row, "name", "이름")
    if not name:
        return None
    grade = _row_int(row, "grade", "학년", default=0)
    class_num = _row_int(row, "class_num", "반", default=0)
    number = _row_int(row, "number", "번호", default=0)
    if grade <= 0 or class_num <= 0 or number <= 0:
        return None

    target = student_identity_key(name=name, grade=grade, class_num=class_num, number=number)
    for student in list_students():
        if student_identity_key(
            name=student.name,
            grade=student.grade,
            class_num=student.class_num,
            number=student.number,
        ) == target:
            return student
    return None


def _finalize_subject_info(info: dict[str, object]) -> dict[str, object]:
    content = str(info.get("content") or info.get("notes") or "").strip()
    activities = info.get("activities")
    if not content and isinstance(activities, list) and activities:
        content = ";".join(str(item).strip() for item in activities if str(item).strip())
    if content and not activities:
        info["activities"] = [content]
        info["notes"] = content
    elif content:
        info["notes"] = content
    return info


def student_from_row(row: dict[str, str]) -> StudentInput:
    student_id = _row_str(row, "id") or new_id()
    subjects: dict[str, dict[str, object]] = {}
    changche: dict[str, str] = {}
    notes: dict[str, object] = {}

    for key, value in row.items():
        if not value or key in {"id", "name", "이름", "grade", "학년", "class_num", "반", "number", "번호", "gender", "성별", "status"}:
            continue
        if key.startswith("세특_"):
            rest = key[3:]
            matched = False
            for suffix, field in REGISTRY_SUBJECT_SUFFIXES:
                if rest.endswith(suffix):
                    subject = rest[: -len(suffix)]
                    if subject:
                        bucket = subjects.setdefault(subject, {})
                        bucket[field] = value
                    matched = True
                    break
            if not matched:
                subjects[rest] = _finalize_subject_info(
                    {
                        "activities": _split_multi(value),
                        "notes": value,
                    }
                )
        elif key.startswith("subject_"):
            subject = key.split("_", 1)[1]
            subjects[subject] = _finalize_subject_info(
                {
                    "activities": _split_multi(value),
                    "notes": row.get(f"{subject}_notes", ""),
                }
            )
        elif key.startswith("창체_") or key.startswith("changche_"):
            subsection = key.split("_", 1)[1]
            if subsection != "봉사":
                changche[subsection] = value
        elif key in ("행발_notes", "행발_메모", "행발_keywords", "행발_키워드", "notes"):
            notes[key] = value
        elif key.endswith("_activities"):
            subject = key.replace("_activities", "")
            subjects.setdefault(subject, {})["activities"] = _split_multi(value)

    for subject, info in list(subjects.items()):
        subjects[subject] = _finalize_subject_info(info)

    keywords_raw = _row_str(row, "행발_keywords", "행발_키워드")
    if "keywords" not in notes and keywords_raw:
        notes["keywords"] = _split_multi(keywords_raw)
    write_targets = _write_targets_from_row(row)
    if write_targets:
        notes["write_targets"] = write_targets

    haengbal = _row_str(row, "행발_notes", "행발_메모", "notes")

    return StudentInput(
        id=student_id,
        name=_row_str(row, "name", "이름"),
        grade=_row_int(row, "grade", "학년", default=1) or 1,
        class_num=_row_int(row, "class_num", "반", default=1) or 1,
        number=_row_int(row, "number", "번호", default=1) or 1,
        gender=_row_str(row, "gender", "성별"),
        status=_row_str(row, "status") or "pending",
        notes={
            "행발": haengbal,
            "keywords": notes.get("keywords", []),
            **{k: v for k, v in notes.items() if k not in {"keywords"}},
        },
        subjects=subjects,
        changche={
            "자율": changche.get("자율", _row_str(row, "창체_자율")),
            "동아리": changche.get("동아리", _row_str(row, "창체_동아리")),
            "진로": changche.get("진로", _row_str(row, "창체_진로")),
        },
    )


def _collect_registry_subjects(students: list[StudentInput]) -> list[str]:
    seen: set[str] = set()
    subjects: list[str] = []
    for student in students:
        for subject in student.subjects.keys():
            if subject not in seen:
                seen.add(subject)
                subjects.append(subject)
    return subjects


def registry_headers(students: list[StudentInput]) -> list[str]:
    subjects = _collect_registry_subjects(students)
    headers = [
        "id",
        "name",
        "grade",
        "class_num",
        "number",
        "gender",
        "write_targets",
        "행발_notes",
        "행발_keywords",
    ]
    for subject in subjects:
        headers.extend(
            [
                f"세특_{subject}_진로",
                f"세특_{subject}_수행평가",
                f"세특_{subject}_주제",
                f"세특_{subject}_활동",
            ]
        )
    headers.extend(f"창체_{key}" for key in REGISTRY_CHANGCHE_KEYS)
    return headers


def student_to_registry_row(student: StudentInput, subjects: list[str]) -> dict[str, str]:
    notes = student.notes or {}
    keywords = notes.get("keywords") or []
    if isinstance(keywords, list):
        keyword_text = "|".join(str(item).strip() for item in keywords if str(item).strip())
    else:
        keyword_text = str(keywords).strip()

    row: dict[str, str] = {
        "id": student.id,
        "name": student.name,
        "grade": str(student.grade),
        "class_num": str(student.class_num),
        "number": str(student.number),
        "gender": student.gender or "",
        "write_targets": "|".join(
            str(item).strip() for item in notes.get("write_targets", []) if str(item).strip()
        )
        if isinstance(notes.get("write_targets"), list)
        else str(notes.get("write_targets") or ""),
        "행발_notes": str(notes.get("행발") or notes.get("행발_notes") or ""),
        "행발_keywords": keyword_text,
    }

    for subject in subjects:
        info = student.subjects.get(subject) or {}
        content = str(info.get("content") or info.get("notes") or "").strip()
        if not content:
            activities = info.get("activities")
            if isinstance(activities, list):
                content = ";".join(str(item).strip() for item in activities if str(item).strip())
        row[f"세특_{subject}_진로"] = str(info.get("career") or "")
        row[f"세특_{subject}_수행평가"] = str(info.get("assessment_type") or "")
        row[f"세특_{subject}_주제"] = str(info.get("topic") or "")
        row[f"세특_{subject}_활동"] = content

    changche = student.changche or {}
    for key in REGISTRY_CHANGCHE_KEYS:
        row[f"창체_{key}"] = str(changche.get(key) or "")
    return row


def read_registry_table(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_cells = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(cell or "").strip() for cell in header_cells]
        rows: list[dict[str, str]] = []
        for values in rows_iter:
            row = {
                headers[index]: str((values[index] if index < len(values) else "") or "").strip()
                for index in range(len(headers))
                if headers[index]
            }
            if any(row.values()):
                rows.append(row)
        return rows
    return read_table_file(path)


def export_students_registry_tsv(students: list[StudentInput] | None = None) -> bytes:
    rows = students if students is not None else list_students()
    subjects = _collect_registry_subjects(rows)
    headers = registry_headers(rows)
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, delimiter="\t", lineterminator="\n")
    writer.writeheader()
    for student in rows:
        writer.writerow(student_to_registry_row(student, subjects))
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def export_students_registry_xlsx(students: list[StudentInput] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = students if students is not None else list_students()
    subjects = _collect_registry_subjects(rows)
    headers = registry_headers(rows)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "학생목록"
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for student in rows:
        row = student_to_registry_row(student, subjects)
        worksheet.append([row.get(header, "") for header in headers])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def preview_students_import(path: Path) -> dict[str, object]:
    rows = read_registry_table(path)
    duplicates: list[dict[str, object]] = []
    errors: list[str] = []

    for index, row in enumerate(rows, start=1):
        name = _row_str(row, "name", "이름")
        if not name:
            errors.append(f"{index}행: 이름이 비어 있습니다.")
            continue
        try:
            student_from_row(row)
        except (TypeError, ValueError) as exc:
            errors.append(f"{index}행: {exc}")
            continue
        existing = find_existing_for_row(row)
        if existing:
            duplicates.append(
                {
                    "row_index": index,
                    "name": name,
                    "grade": _row_int(row, "grade", "학년", default=1),
                    "class_num": _row_int(row, "class_num", "반", default=1),
                    "number": _row_int(row, "number", "번호", default=1),
                    "existing_id": existing.id,
                }
            )

    valid_rows = len(rows) - len(errors)
    return {
        "total": len(rows),
        "valid_count": valid_rows,
        "new_count": valid_rows - len(duplicates),
        "duplicate_count": len(duplicates),
        "duplicates": duplicates,
        "errors": errors,
    }


def merge_student_from_row(existing: StudentInput, row: dict[str, str]) -> StudentInput:
    incoming = student_from_row({**row, "id": existing.id})
    if _row_has_key(row, "name", "이름"):
        existing.name = incoming.name
    if _row_has_key(row, "grade", "학년"):
        existing.grade = incoming.grade
    if _row_has_key(row, "class_num", "반"):
        existing.class_num = incoming.class_num
    if _row_has_key(row, "number", "번호"):
        existing.number = incoming.number
    if _row_has_key(row, "gender", "성별"):
        existing.gender = incoming.gender

    notes = dict(existing.notes or {})
    if _row_has_key(row, "행발_notes", "행발_메모", "notes"):
        notes["행발"] = _row_str(row, "행발_notes", "행발_메모", "notes")
    if _row_has_key(row, "행발_keywords", "행발_키워드"):
        notes["keywords"] = _split_multi(_row_str(row, "행발_keywords", "행발_키워드"))
    if _row_has_key(row, *REGISTRY_WRITE_TARGET_KEYS):
        notes["write_targets"] = _write_targets_from_row(row)
    existing.notes = notes

    subjects = dict(existing.subjects or {})
    for subject in _subjects_present_in_row(row):
        if subject in incoming.subjects:
            subjects[subject] = incoming.subjects[subject]
        else:
            subjects.pop(subject, None)
    existing.subjects = subjects

    changche = dict(existing.changche or {})
    for key in REGISTRY_CHANGCHE_KEYS:
        if _row_has_key(row, f"창체_{key}", f"changche_{key}"):
            changche[key] = _row_str(row, f"창체_{key}", f"changche_{key}")
    existing.changche = changche
    return save_student(existing)


def import_students_registry(path: Path, mode: str = "add") -> dict[str, object]:
    if mode not in {"add", "skip", "update"}:
        raise ValueError("mode는 add, skip, update 중 하나여야 합니다.")

    rows = read_registry_table(path)
    imported: list[str] = []
    updated: list[str] = []
    skipped: list[str] = []
    errors: list[str] = []

    for index, row in enumerate(rows, start=1):
        name = _row_str(row, "name", "이름")
        if not name:
            errors.append(f"{index}행: 이름이 비어 있습니다.")
            continue
        try:
            existing = find_existing_for_row(row)
            if existing:
                if mode == "skip":
                    skipped.append(existing.id)
                    continue
                if mode == "update":
                    updated.append(merge_student_from_row(existing, row).id)
                    continue
            imported.append(add_student(student_from_row(row)).id)
        except (TypeError, ValueError) as exc:
            errors.append(f"{index}행: {exc}")

    students = [get_student(student_id) for student_id in imported + updated]
    students = [student for student in students if student is not None]
    return {
        "imported": len(imported),
        "updated": len(updated),
        "skipped": len(skipped),
        "errors": errors,
        "students": [student.to_dict() for student in students],
    }


def import_students_file(path: Path, mode: str = "add") -> list[StudentInput]:
    result = import_students_registry(path, mode=mode)
    students: list[StudentInput] = []
    for data in result.get("students", []):
        if isinstance(data, dict):
            students.append(StudentInput.from_dict(data))
    return students
