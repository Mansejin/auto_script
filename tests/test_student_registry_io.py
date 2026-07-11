from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.saenggibu.models import StudentInput
from src.saenggibu.student_store import (
    add_student,
    export_students_registry_tsv,
    export_students_registry_xlsx,
    find_existing_for_row,
    import_students_registry,
    list_students,
    preview_students_import,
    read_registry_table,
)


@pytest.fixture()
def students_dir(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> Path:
    root = tmp_path / "data" / "saenggibu" / "students"
    outputs = tmp_path / "data" / "saenggibu" / "outputs"
    root.mkdir(parents=True)
    outputs.mkdir(parents=True)
    monkeypatch.setattr("src.saenggibu.student_store.STUDENTS_DIR", root)
    monkeypatch.setattr("src.saenggibu.student_store.OUTPUTS_DIR", outputs)
    monkeypatch.setenv("SGB_STORE_GENERATED", "1")
    monkeypatch.setenv("SGB_ENCRYPT_DATA", "0")
    return root


def test_export_and_import_registry_roundtrip(students_dir: Path) -> None:
    add_student(
        StudentInput(
            id="sround01",
            name="김민수",
            grade=2,
            class_num=3,
            number=5,
            notes={"행발": "행발 메모", "keywords": ["책임감", "성실"]},
            subjects={
                "국어": {
                    "career": "인문",
                    "assessment_type": "보고서",
                    "topic": "토론",
                    "content": "토론 수업 참여",
                    "activities": ["토론 수업 참여"],
                    "notes": "토론 수업 참여",
                }
            },
            changche={"자율": "학급 환경 미화", "동아리": "과학 동아리", "진로": "공학 탐색"},
            generated={"행발": "작성본"},
            status="done",
        )
    )

    tsv_path = students_dir / "export.tsv"
    tsv_path.write_bytes(export_students_registry_tsv())
    rows = read_registry_table(tsv_path)
    assert rows[0]["name"] == "김민수"
    assert rows[0]["행발_notes"] == "행발 메모"
    assert rows[0]["세특_국어_활동"] == "토론 수업 참여"
    assert "창체_봉사" not in rows[0]

    preview = preview_students_import(tsv_path)
    assert preview["duplicate_count"] == 1
    assert preview["new_count"] == 0

    result = import_students_registry(tsv_path, mode="update")
    assert result["updated"] == 1
    updated = list_students()[0]
    assert updated.generated.get("행발") == "작성본"
    assert updated.status == "done"
    assert updated.notes["행발"] == "행발 메모"


def test_import_skip_duplicate(students_dir: Path) -> None:
    add_student(StudentInput(id="", name="이서연", grade=2, class_num=1, number=1))
    tsv_path = students_dir / "dup.tsv"
    tsv_path.write_text(
        "name\tgrade\tclass_num\tnumber\t행발_notes\n이서연\t2\t1\t1\t새 메모\n",
        encoding="utf-8",
    )
    result = import_students_registry(tsv_path, mode="skip")
    assert result["skipped"] == 1
    assert list_students()[0].notes.get("행발") != "새 메모"


def test_export_registry_xlsx(students_dir: Path) -> None:
    add_student(
        StudentInput(
            id="",
            name="박지훈",
            grade=1,
            class_num=2,
            number=7,
            notes={"행발": "메모"},
        )
    )
    workbook = load_workbook(BytesIO(export_students_registry_xlsx()))
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    assert "행발_notes" in headers
    assert "창체_봉사" not in headers


def test_find_existing_for_row_by_identity(students_dir: Path) -> None:
    student = add_student(StudentInput(id="keep-id", name="최유진", grade=2, class_num=4, number=8))
    row = {"name": "최유진", "grade": "2", "class_num": "4", "number": "8"}
    existing = find_existing_for_row(row)
    assert existing is not None
    assert existing.id == student.id


def test_exported_tsv_id_roundtrips_with_bom_after_identity_edit(students_dir: Path) -> None:
    add_student(StudentInput(id="rename-id", name="정하늘", grade=1, class_num=2, number=3))

    tsv_text = export_students_registry_tsv().decode("utf-8")
    edited = tsv_text.replace("정하늘\t1\t2\t3", "정하늘수정\t1\t2\t3")
    tsv_path = students_dir / "rename.tsv"
    tsv_path.write_text(edited, encoding="utf-8")

    rows = read_registry_table(tsv_path)
    assert rows[0]["id"] == "rename-id"

    result = import_students_registry(tsv_path, mode="update")
    assert result["updated"] == 1
    students = list_students()
    assert len(students) == 1
    assert students[0].id == "rename-id"
    assert students[0].name == "정하늘수정"


def test_update_rejects_id_with_another_existing_student_identity(students_dir: Path) -> None:
    add_student(
        StudentInput(
            id="student-a",
            name="학생A",
            grade=1,
            class_num=1,
            number=1,
            notes={"행발": "A 원본"},
            generated={"행발": "A 작성본"},
            status="done",
        )
    )
    add_student(
        StudentInput(
            id="student-b",
            name="학생B",
            grade=1,
            class_num=1,
            number=2,
            notes={"행발": "B 원본"},
        )
    )
    tsv_path = students_dir / "conflict.tsv"
    tsv_path.write_text(
        "id\tname\tgrade\tclass_num\tnumber\t행발_notes\n"
        "student-a\t학생B\t1\t1\t2\tB 수정 메모\n",
        encoding="utf-8",
    )

    result = import_students_registry(tsv_path, mode="update")
    assert result["updated"] == 0
    assert result["errors"]

    students = {student.id: student for student in list_students()}
    assert students["student-a"].name == "학생A"
    assert students["student-a"].notes["행발"] == "A 원본"
    assert students["student-a"].generated["행발"] == "A 작성본"
    assert students["student-a"].status == "done"
    assert students["student-b"].notes["행발"] == "B 원본"


def test_read_registry_xlsx_handles_rows_shorter_than_headers(students_dir: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["id", "name", "grade", "class_num", "number", "행발_notes"])
    worksheet.append(["short-id", "김단축"])
    xlsx_path = students_dir / "short.xlsx"
    workbook.save(xlsx_path)

    rows = read_registry_table(xlsx_path)
    assert rows == [
        {
            "id": "short-id",
            "name": "김단축",
            "grade": "",
            "class_num": "",
            "number": "",
            "행발_notes": "",
        }
    ]
