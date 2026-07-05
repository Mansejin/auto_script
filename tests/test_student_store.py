from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.saenggibu.models import StudentInput
from src.saenggibu.student_store import (
    add_student,
    delete_all_students,
    delete_student,
    delete_students,
    export_students_xlsx,
    get_student,
    list_students,
    reconcile_students,
    reset_all_generated,
    reset_generated_for_students,
)


@pytest.fixture()
def students_dir(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> Path:
    root = tmp_path / "data" / "saenggibu" / "students"
    outputs = tmp_path / "data" / "saenggibu" / "outputs"
    root.mkdir(parents=True)
    outputs.mkdir(parents=True)
    monkeypatch.setattr("src.saenggibu.student_store.STUDENTS_DIR", root)
    monkeypatch.setattr("src.saenggibu.student_store.OUTPUTS_DIR", outputs)
    return root


def test_delete_student_removes_json(students_dir: Path) -> None:
    student = add_student(
        StudentInput(id="", name="김민수", grade=2, class_num=1, number=1, notes={"행발": "메모"})
    )
    assert delete_student(student.id) is True
    assert get_student(student.id) is None
    assert delete_student(student.id) is False


def test_delete_students_bulk(students_dir: Path) -> None:
    a = add_student(StudentInput(id="", name="A", grade=1, class_num=1, number=1))
    b = add_student(StudentInput(id="", name="B", grade=1, class_num=1, number=2))
    result = delete_students([a.id, b.id, "missing"])
    assert set(result["deleted"]) == {a.id, b.id}
    assert result["not_found"] == ["missing"]
    assert list_students() == []


def test_delete_all_students(students_dir: Path) -> None:
    add_student(StudentInput(id="", name="A", grade=1, class_num=1, number=1))
    add_student(StudentInput(id="", name="B", grade=1, class_num=1, number=2))
    assert delete_all_students() == 2
    assert list_students() == []


def test_reset_generated_keeps_student(students_dir: Path) -> None:
    student = add_student(
        StudentInput(
            id="",
            name="김민수",
            grade=2,
            class_num=1,
            number=1,
            notes={"행발": "메모"},
            generated={"행발": "작성됨"},
            status="done",
        )
    )
    result = reset_generated_for_students([student.id])
    assert result["count"] == 1
    updated = get_student(student.id)
    assert updated is not None
    assert updated.generated == {}
    assert updated.status == "pending"
    assert updated.notes["행발"] == "메모"


def test_reset_all_generated(students_dir: Path) -> None:
    add_student(
        StudentInput(
            id="",
            name="A",
            grade=1,
            class_num=1,
            number=1,
            generated={"행발": "본문"},
            status="done",
        )
    )
    add_student(StudentInput(id="", name="B", grade=1, class_num=1, number=2))
    assert reset_all_generated() == 1


def test_export_students_xlsx(students_dir: Path) -> None:
    add_student(
        StudentInput(
            id="",
            name="김민수",
            grade=2,
            class_num=3,
            number=12,
            generated={"행발": "행발 본문", "세특": {"윤사": "세특 본문"}, "창체": {"자율": "자율 본문"}},
            status="done",
        )
    )
    content = export_students_xlsx()
    from io import BytesIO

    wb = load_workbook(BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "행발" in headers
    assert "세특_윤사" in headers
    row = [cell.value for cell in ws[2]]
    assert row[headers.index("이름")] == "김민수"
    assert row[headers.index("행발")] == "행발 본문"
    assert row[headers.index("세특_윤사")] == "세특 본문"
    assert row[headers.index("자율")] == "자율 본문"


def test_delete_student_when_json_id_differs_from_filename(students_dir: Path) -> None:
    path = students_dir / "sddb4b24a.json"
    path.write_text(
        '{"name":"김민수","grade":2,"class_num":3,"number":12,"id":"swrong0001"}',
        encoding="utf-8",
    )
    listed = list_students()
    assert len(listed) == 1
    assert listed[0].id == "sddb4b24a"
    assert listed[0].name == "김민수"
    assert delete_student(listed[0].id) is True
    assert list(students_dir.glob("*.json")) == []


def test_reconcile_removes_ghost_student(students_dir: Path) -> None:
    ghost = students_dir / "sghost0001.json"
    ghost.write_text("{}", encoding="utf-8")
    add_student(StudentInput(id="", name="정상", grade=2, class_num=1, number=3))
    result = reconcile_students()
    assert "sghost0001.json" in result["removed"]
    students = list_students()
    assert len(students) == 1
    assert students[0].name == "정상"
